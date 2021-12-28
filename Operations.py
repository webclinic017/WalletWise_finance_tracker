# import csv
import csv

# import pandas
import pandas as pd

# import numpy
import numpy as np

# import openpyxl
import openpyxl as xl

# import workbook from openpyxl
from openpyxl import load_workbook

# import tkinter library
from tkinter import *

# import partial from functools
from functools import partial

# import re liibrary
import re


# import testingproject1_calc

# create class for user identification
class Identity():

    # create a constructor
    def __init__(self):
        self.username = ""
        self.password = ""
        self.email = ""

    # create a method for users details
    def sign_up(username, password, email):
        filename = "user_details.csv"
        data = [[username, password, email]]

        # write in csv file
        with open(filename, 'a', newline='') as csvfile:
            # create a csv dictionary writer object
            writer = csv.writer(csvfile)

            # write data rows
            writer.writerows(data)

    # create a method for users list of usernames
    def username_list():
        username_list = []
        with open('user_details.csv', 'r') as f:
            for line in f:
                username_list.append(line.split(',')[0])
            return username_list

    # create a method to list users usernames, password, and email
    def user_matrix(username, password, no):
        password_list = []
        email_list = []
        username_list = []
        with open('user_details.csv', 'r') as f:
            for line in f:
                password_list.append(line.split(',')[1])
            for line in f:
                email_list.append(line.split(',')[2])
            for line in f:
                username_list.append(line.split(',')[0])

        data_password = password_list[no]
        while password != data_password:
            print('Incorrect password')

            while True:
                try:
                    user_input = input("Type 1 to retry and 0 to exit: ")
                    if user_input != "1" and user_input != "0":
                        raise ValueError
                    break

                except ValueError:
                    print("Invalid input, the system expected 0 or 1")

            if user_input == "1":
                password = input("Enter Password: ")
            if user_input == "0":
                break
        if password == data_password:
            print("Login Successful!")
            return True

        else:
            print("Invalid input \n")

    # create a method to connect users' details to spreadsheet
    def new_excel_sheet(username):

        # calling the spreadsheet
        path = ("user_db.xlsx")
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book

        expenses_sheet = np.random.randn(0, 0)
        df_expenses = pd.DataFrame(expenses_sheet)

        income_sheet = np.random.randn(0, 0)
        df_income = pd.DataFrame(income_sheet)

        # name new excel sheets with usernames
        df_expenses.to_excel(writer, sheet_name=f'{username}')
        df_income.to_excel(writer, sheet_name=f'{username}_income')

        # save and close sheets
        writer.save()
        writer.close()

        # copy data from first (template) sheet to the new created sheet
        wb1 = xl.load_workbook(path)

        # open two destination sheets to record new user's expenses and income
        ws1 = wb1.worksheets[0]  # First sheet to be copied
        ws3 = wb1.worksheets[1]  # Sheet two: to be copied:template for income
        ws2 = wb1.worksheets[-2]  # second last sheet
        ws4 = wb1.worksheets[-1]  # last sheet

        # calculate total number of rows and columns in template sheets
        tr = ws3.max_row
        tc = ws1.max_column
        tr1 = ws3.max_row
        tc1 = ws3.max_column

        # copy the cell values from template sheet for expenses
        for i in range(1, tr + 1):
            for j in range(1, tc + 1):
                # reading cell value from template excel sheet
                c = ws1.cell(row=i, column=j)

                # writing the read value to destination excel sheet
                ws2.cell(row=i, column=j).value = c.value

        # copying the cell values from the template sheet for income
        for i in range(1, tr1 + 1):
            for j in range(1, tc1 + 1):
                # reading cell value from template excel sheet
                c = ws3.cell(row=i, column=j)

                # writing the read value to destination excel sheet
                ws4.cell(row=i, column=j).value = c.value

        # saving the excel file
        # wb1.save(str(path))
        wb1.save("user_db.xlsx")
        # return path

        # return username

    # create a method for better user interface
    def validateLogin(username, password):
        print("Username entered :", username.get())
        print("Password entered :", password.get())
        return

    # window
    # tkWindow = Tk()
    # tkWindow.geometry('400x150')
    # tkWindow.title('Tkinter Login Form - pythonexamples.org')

    # # username label and text entry box
    # usernameLabel = Label(tkWindow, text="User Name").grid(row=0, column=0)
    # username = StringVar()
    # usernameEntry = Entry(tkWindow, textvariable=username).grid(row=0, column=1)

    # # password label and password entry box
    # passwordLabel = Label(tkWindow, text="Password").grid(row=1, column=0)
    # password = StringVar()
    # passwordEntry = Entry(tkWindow, textvariable=password, show='*').grid(row=1, column=1)

    # validateLogin = partial(validateLogin, username, password)

    # # login button
    # loginButton = Button(tkWindow, text="Login", command=validateLogin).grid(row=4, column=0)

    # tkWindow.mainloop()


# create a function to verify the password
def Check_password():
    password = ""
    while password == "":
        password = input('Enter your password: ')

        # checking password length
        if len(password) >= 8:
            if (bool(re.match('((?=.*\d)(?=.*[a-z])(?=.*[A-Z])(?=.*[!@#$%^&*])(?=.*[0-9]))', password)) == True):
                continue

            # checking if the password has uppercase letters
            if re.search(r"[A-Z]", password) is None:
                print('Your password is weak, please consider including at least one UPPERCASE letter')
                password = ""

            # checking if the password has lowercase letters
            if re.search(r"[a-z]", password) is None:
                print("Weak password, please consider including at least one lowercase letter")
                password = ""

            # checking if the password contains special characters
            if re.search(r"[!@#$%^&*]", password) is None:
                print("Weak password, please consider including at least one special character; !@#$%^&*")
                password = ""

            # checking if the password contains at least one number
            if re.search(r"[0-9]", password) is None:
                print("Weak password, please consider including at least one number")
                password = ""

        else:
            print(
                '''Please enter a valid password.\nPassword should contain at least 8 characters, one upper and lower case letters, a number, and a special character''')
            password = ""

    return password


def Check_password_testing():
    password = input('Enter your password: ')

    # checking password length
    if len(password) >= 8:
        if (bool(re.match('((?=.*\d)(?=.*[a-z])(?=.*[A-Z])(?=.*[!@#$%^&*])(?=.*[0-9]))', password)) == True):
            return "Valid password"

        # checking if the password has uppercase letters
        if re.search(r"[A-Z]", password) is None:
            return 'Your password is weak, please consider including at least one UPPERCASE letter'

        # checking if the password has lowercase letters
        if re.search(r"[a-z]", password) is None:
            return "Weak password, please consider including at least one lowercase letter"

        # checking if the password contains special characters
        if re.search(r"[!@#$%^&*]", password) is None:
            return "Weak password, please consider including at least one special character; !@#$%^&*"

        # checking if the password contains at least one number
        if re.search(r"[0-9]", password) is None:
            return "Weak password, please consider including at least one number"


    else:
        return '''Please enter a valid password.\nPassword should contain at least 8 characters, one upper and lower case letters, a number, and a special character'''

    # return password


# create function to validate emails
def Check_email():
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email = ""
    while email == "":
        email = input("Enter your email: ")
        if (re.fullmatch(regex, email)):
            continue
        else:
            print("Invalid Email, Please enter a correct email address")
            email = ""
    return email


def Check_email_testing():
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

    email = input("Enter your email: ")
    if (re.fullmatch(regex, email)):
        return "Valid Email"
    else:
        return "Invalid Email, Please enter a correct email address"

