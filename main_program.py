import openpyxl
from openpyxl.worksheet import worksheet
import time, sys
from tkinter import *
from tkcalendar import Calendar
import calcc
import sendmail
import pandas as pd


# creating a parent class Account
class Account:
    def __init__(self, balance):
        self.balance = balance

    # creating a method to record income in the excel sheet
    def earnings(self, amount_earned):

        if amount_earned < 0:
            print("Please a valid amount. Positive integers only")
            raise ValueError

        else:
            self.balance += amount_earned

            return self.balance

    # creating a method to record expenses in the excel sheet
    def spend(self, amount_spent):

        if amount_spent < 0:
            print("Please a valid amount. Positive integers only")
            raise ValueError

        else:
            self.balance -= amount_spent
            print("New balance is:", self.balance)
            if self.balance < 0:
                print("You are going over your budget! Please be cautious.")
            return self.balance

    # creating a method to transfer money between different accounts
    def transfer(self, amount, from_ac, to_ac):

        print("from", from_ac.balance)
        print("to", to_ac.balance)

        # Check if the transfer is to happen between the same account, if so, raise a valueError
        if from_ac == to_ac:
            raise ValueError

        # Check if the amount to be transferred exceeds the amount in the account
        if from_ac.balance < amount:
            raise ValueError

        output = []
        output.append(from_ac.balance - amount)
        output.append(to_ac.balance + amount)

        return output


# creating Momo class to inherit form Account
class Momo(Account):
    def __init__(self, balance):
        super().__init__(balance)
        self.balance = balance


# creating Bank class to inherit form Account
class Bank(Account):
    def __init__(self, balance):
        super().__init__(balance)
        self.balance = balance


# creating Cash class to inherit form Account
class Cash(Account):
    def __init__(self, balance):
        super().__init__(balance)
        self.balance = balance


# creating a method to decrement from the correct account based on a users spending
def checkaccount(userinput, price):
    # Check the user account selected
    if userinput.lower() == "momo":

        resultant = momo_ac.spend(price)
        sheet2.cell(2, 1).value = resultant
        wb_obj.save("user_db.xlsx")


    # Check the user account selected
    elif userinput.lower() == "bank":

        resultant = bank_ac.spend(price)
        sheet2.cell(2, 3).value = resultant
        wb_obj.save("user_db.xlsx")

    # Check the user account selected
    elif userinput.lower() == "cash":

        resultant = cash_ac.spend(price)
        sheet2.cell(2, 2).value = resultant
        wb_obj.save("user_db.xlsx")


# creating a function to record users expenses
def recordExpense():
    # declaring a variable for the category of expense
    expense_categories = {1: "Food", 2: "Transport", 3: "Rent", 4: "Bills", 5: "Shopping/Groceries", 6: "Gifts",
                          7: "Travel", 8: "Healthcare", 9: "Clothes/Shoes", 10: "Others"}

    # declaring a variable for the accounts (for the user payment model)
    income_categories = {1: "Momo", 2: "Bank", 3: "Cash"}

    # Creating an empty list to store the dates
    # date_lists = []
    print(date_lists)
    # Looping through to append the dates into the list
    # for cell in range(1, rows1 + 1):
    #     date_lists.append(sheet1.cell(cell, 1).value)

    while True:
        # Ask user which category of items they spent on
        try:

            category = int(input("""Which category of items did you spend on ?
                                1 -> Food
                                2 -> Transport
                                3 -> Rent
                                4 -> Bills
                                5 -> Shopping/Groceries
                                6 -> Gifts
                                7 -> Travel
                                8 -> Healthcare
                                9 -> Clothes/Shoes
                                10 -> Others \n"""))
            # Check if user input doesn't match the options available, if so, raise error
            if category < 1 or category > 10:
                raise ValueError
            break

        # Alert the user that their input is invalid
        except ValueError:
            print("Please enter a valid entry. ")

    while True:
        # Ask user how much they spent of the chosen category of items
        try:
            price = int(input("How much did you spend on " + expense_categories[category].lower() + " ?\n"))
            # Raise error if a negative price is entered
            if price < 0:
                raise ValueError
            break

        # Alert the user that their input is wrong
        except ValueError:
            print("Please a valid amount. Positive integers only ")

    while True:
        # Ask user which account they used
        try:
            ac = int(input("""What account did you use? 
                            1 -> Momo
                            2 -> Bank
                            3 -> Cash \n"""))

            # Raise error if wrong input is entered
            if ac < 1 or ac > 3:
                raise ValueError
            break

        # Alert the user that their input is wrong
        except ValueError:
            print("Please enter a valid entry. ")

    print("When was the transaction made?")
    time.sleep(3)

    # calling calendar for the user to choose a date
    thedate = calcc.mycal()
    print(thedate)
    # Check if the inputted date is in the existing list of dates
    if thedate not in date_lists:
        # Append the date into the list if it does not exist
        date_lists.append(thedate)

        # Insert the date into the next empty cell in the date column

        sheet1.cell(rows1 + 1, 1).value = thedate
        wb_obj.save("user_db.xlsx")

        # wb_obj.save("user_db.xlsx")

    # receive the balance from the account we are crediting from
    bal = checkaccount(income_categories[ac], price)

    # Loop through the cells to input the new balance
    for j in range(1, cols1 + 1):

        if sheet1.cell(1, j).value.lower() == expense_categories[category].lower():
            print(date_lists.index(thedate))
            c = sheet1.cell(date_lists.index(thedate) + 1, j)

            # c.value = bal
            # If the cell is not empty, add it to the existing balance
            if c.value != None:
                c.value += price
                wb_obj.save("user_db.xlsx")
            # # If the cell is empty, the value of the cell becomes the inoutted balance
            else:
                c.value = price
                wb_obj.save("user_db.xlsx")
            print("Your expense has successfully been recorded")
            # print("You've spent:", c.value)


# creating a function to record users' income
def recordIncome():
    while True:
        # Ask user which account they want to add an income tp
        try:
            account = int(input("""Which account balance would you like to increment? 
                                1 -> Momo
                                2 -> Bank
                                3 -> Cash\n"""))

            # Check if the user inputs a valid entry, if not, raise ValueError
            if account < 1 or account > 3:
                raise ValueError

            break

        # Alert the user to make a valid entry
        except ValueError:
            print("Kindly enter a valid entry. \n")

    while True:
        # Ask user how much they want to record
        try:
            amount = int(input("How much would you like to record? \n"))

            # Check if the user inputs a valid entry, if not, raise ValueError
            if amount < 0:
                raise ValueError
            break

        # Alert the user to make a valid entry
        except ValueError:
            print("Kindly enter amount in figures. \n")

    # Check the account that the user entered, pass this account to the earnings method in order to increment the right account
    # Increment momo account
    if account == 1:
        result = momo_ac.earnings(amount)
        sheet2.cell(2, 1).value = result
        wb_obj.save("user_db.xlsx")

    # Increment bank account
    elif account == 2:
        result = bank_ac.earnings(amount)
        sheet2.cell(2, 3).value = result
        wb_obj.save("user_db.xlsx")

    # Increment cash account
    elif account == 3:
        result = cash_ac.earnings(amount)
        sheet2.cell(2, 2).value = result
        wb_obj.save("user_db.xlsx")

    print("Your income has successfully been recorded.")
    return


# creating a function that allows users to transfer money among different accounts
def transferMoney():
    momo_ac_bal = momo_ac.balance
    bank_ac_bal = bank_ac.balance
    cash_ac_bal = cash_ac.balance
    # Dictionary to hold the different account names
    account_dict = {1: momo_ac, 2: bank_ac, 3: cash_ac}

    # Print the existing account types and corresponding balances
    print("""Your current account balances are as follows:
            -> Momo: %d rwf
            -> Bank: %d rwf
            -> Cash: %d rwf
            """ % (momo_ac_bal, bank_ac_bal, cash_ac_bal))

    while True:
        # Ask user which account they would like to transfer funds from
        try:
            transfer_from = int(input("""Which account would you like to transfer funds from? 
                                1 -> Momo
                                2 -> Bank
                                3 -> Cash \n"""))

            # Check if user input meets requirements, if not, raise ValueError
            if transfer_from < 1 or transfer_from > 3:
                raise ValueError
            break

        # Prompt user to enter a valid entry
        except ValueError:
            print("Please enter a valid entry. ")

    while True:
        # Ask user which account they would like to transfer funds to
        try:
            transfer_to = int(input("""Which account would you like to transfer funds from? 
                                1 -> Momo
                                2 -> Bank
                                3 -> Cash \n"""))
            # Check if inout is valid, if not, raise Value Error
            if transfer_to < 1 or transfer_to > 3:
                raise ValueError

            # Confirm that the accounts are not equal, if they are, raise ValueError
            if transfer_from == transfer_to:
                raise ValueError

            break

        # Alert user to enter valid entries
        except ValueError:
            print("Please enter a valid entry. Please not that you can't transfer to and from the same account")

    while True:
        # Ask user amount they would like to transfer
        try:
            transfer_amount = int(input("How much money would you like to transfer between the accounts? \n"))
            # raise ValueError if account balance of account from is less than the amount to be sent
            if account_dict[transfer_from].balance < transfer_amount:
                raise ValueError
            break

        # Alert user that they have insufficient balance in their account
        except ValueError:
            print("You have insufficient ammounts in your account. ")

    # If else block to keep track of the cell values of the excel sheet for the account to transfer to and from
    if transfer_to == 1:
        row_to = 2
        col_to = 1

    elif transfer_to == 2:
        row_to = 2
        col_to = 3

    elif transfer_to == 3:
        row_to = 2
        col_to = 2

    if transfer_from == 1:
        row_from = 2
        col_from = 1

    elif transfer_from == 2:
        row_from = 2
        col_from = 3

    elif transfer_from == 3:
        row_from = 2
        col_from = 2

    # list_result hold the return value after calling transfer function
    list_result = account_dict[transfer_from].transfer(transfer_amount, account_dict[transfer_from],
                                                       account_dict[transfer_to])

    # Insert the new values to the respective cells
    sheet2.cell(row_to, col_to).value = list_result[1]
    sheet2.cell(row_from, col_from).value = list_result[0]

    # Save changes to the excel file
    wb_obj.save("user_db.xlsx")


# creating a function to sum up total amount a user's has spent per day
def totalexpensesperday():
    # creating list of dates
    date_lists = []
    for cell in range(2, rows1 + 1):
        date_lists.append(sheet1.cell(cell, 1).value)

    # ask user the date on which they want to check the amount spent
    print("Which day would you like to know the amount of money you spent?")
    thedate = calcc.mycal()
    print(thedate)

    # checking if the date selected is the list of dates
    if thedate not in date_lists:

        # print a message to let the user know if the selected date is not in the list of dates
        print("You did not record any expenses on", thedate)

    # finding the sum of amount spent on all categories on the selected date
    else:
        sum = 0

        # looping through the rows to calculate the total on a particular date
        for i in range(2, rows1 + 1):
            if sheet1.cell(i, 1).value == thedate:
                for cell in range(2, cols1 + 1):
                    if sheet1.cell(i, cell).value != None:
                        sum += sheet1.cell(i, cell).value
                print("You spent a total of", sum, "rwf on", thedate)
            else:
                continue


# create a function to sum up user's expenses per category
def Category_total():
    # calling user's excel sheet in the workbook
    df = pd.read_excel('user_db.xlsx', sheet_name=uname)

    # declaring a variable for the categories of expenses

    while True:
        try:
            category = int(input("""Which category of items did you spend on ?
                                    1 -> Food
                                    2 -> Transport
                                    3 -> Rent
                                    4 -> Bills
                                    5 -> Shopping/Groceries
                                    6 -> Gifts
                                    7 -> Travel
                                    8 -> Healthcare
                                    9 -> Clothes/Shoes
                                    10 -> Others \n"""))
            # printing the amount spent on food when user's input is 1
            if category == 1:
                print("Total amount spent on food: ", df["Food"].sum())

            # printing the amount spent on transport when user's input is 2
            elif category == 2:
                print("Total amount spent on transport: ", df["Transport"].sum())

            # printing the amount spent on rent when user's input is 3
            elif category == 3:
                print("Total amount spent on rent: ", df["Rent"].sum())

            # printing the amount spent on bills when user's input is 4
            elif category == 4:
                print("Total amount spent on bills: ", df["Bills"].sum())

            # printing the amount spent on shopping/groceries when user's input is 5
            elif category == 5:
                print("Total amount spent on shopping/groceries: ", df["Shopping/Groceries"].sum())

            # printing the amount spent on gifts when user's input is 6
            elif category == 6:
                print("Total amount spent on gifts: ", df["Gifts"].sum())

            # printing the amount spent on travel when user's input is 7
            elif category == 7:
                print("Total amount spent on travel: ", df["Travel"].sum())

            # printing the amount spent on healthcare when user's input is 8
            elif category == 8:
                print("Total amount spent on healthcare: ", df["Healthcare"].sum())

            # printing the amount spent on clothes/shoes when user's input is 9
            elif category == 9:
                print("Total amount spent on clothes/shoes: ", df["Clothes/Shoes"].sum())

            # printing the amount spent on others when user's input is 1o
            elif category == 10:
                print("Total amount spent on others: ", df["Others"].sum())

            if category < 1 or category > 10:
                raise ValueError

            break

        except ValueError:
            print("Please enter a valid option")


# creating function to get the total amount spent at any time
def Recorded_expenses():
    # calling the workbook
    df = pd.read_excel('user_db.xlsx')
    df.head()
    df["Total"] = df["Food"] + df["Transport"] + df["Rent"] + df["Bills"] + df["Shopping/Groceries"] + df[
        "Gifts"] + df["Travel"] + df["Healthcare"] + df["Clothes/Shoes"] + df["Others"]

    # print the the sum per category
    print("The total amount you spent so far: ", df["Total"].sum())


# creating a function for the tips on how to save wisely
def tips():
    # calling the workbook with tips
    tips = pd.read_excel('Tips.xlsx')

    # specifying the specific column for tips from the workbook's active sheet
    df = pd.DataFrame(tips, columns=['5 Tips For Spending Money Wisely'])

    # printing the tips
    print(df)


def viewTransactions():
    # Opening an excel sheet using pandas library
    transactions = pd.read_excel('user_db.xlsx', sheet_name=uname)

    # Load the data in the excel sheet into a pandas datafram and display the datafram
    df = pd.DataFrame(transactions,
                      columns=['Date', 'Food', 'Rent', 'Transport', 'Bills', 'Shopping/Groceries', 'Gifts', 'Travel',
                               'Healthcare', 'Clothes/Shoes', 'Others'])
    print(df)


def viewAccountBalances():
    # Opening an excel sheet using pandas library
    transactions = pd.read_excel('user_db.xlsx', sheet_name=uname + '_income')

    # Load the data in the excel sheet into a pandas datafram and display the datafram
    df = pd.DataFrame(transactions, columns=['Momo', 'Cash', 'Bank'])
    print(df)


# creating a function to slowly print a text
def print_slow(str):
    for character in str:
        sys.stdout.write(character)
        sys.stdout.flush()
        time.sleep(0.05)


# creating the main function to run operations
def mainProgram(username):
    # print(email)

    # open the workbook for users details (expenses and incomes)
    path = ("user_db.xlsx")
    global wb_obj
    # Open the excel sheet stored in Path
    wb_obj = openpyxl.load_workbook(path)
    # Save any changes made to the excel sheet
    wb_obj.save("user_db.xlsx")
    # return username

    # Declaring global variables that will be used throughout the program
    global sheet1
    global sheet2
    global momo_ac
    global cash_ac
    global bank_ac
    global cols1
    global rows1
    global uname
    global date_lists

    # saving username as a variable
    uname = username
    date_lists = []

    # assigning a variables user's expense sheet
    sheet1 = wb_obj[username]

    # assigning a variables user's income sheet
    sheet2 = wb_obj[username + "_income"]

    # counting maximum rows and columns from the expenses sheet
    cols1 = sheet1.max_column
    rows1 = sheet1.max_row

    # declaring accounts and their balance on the income sheet
    momo_ac = Momo(sheet2.cell(2, 1).value)
    cash_ac = Cash(sheet2.cell(2, 2).value)
    bank_ac = Bank(sheet2.cell(2, 3).value)

    for cell in range(1, rows1 + 1):
        date_lists.append(sheet1.cell(cell, 1).value)
    # greeting the user after login in
    print_slow("Hello! Welcome to Wallet Wise \n")

    print_slow("Enabling users to have more control and intentionality over their spending :) \n")

    print("How may we help you today? \n")

    tryagain = 1
    # Loop to check if the user would like to perform another operation
    while tryagain == 1:
        while True:
            # Menu that displays list of operations that the user can perform
            try:
                response = int(input("""What operation would you like to be performed? 
                    1 -> Add an expense
                    2 -> Add an income
                    3 -> Transfer money between accounts
                    4 -> Get the total amount you spent on a category of expense
                    5 -> Get the total amount you spent on a specific day
                    6 -> Get the total amount you spent so far
                    7 -> Get tips on how to spend wisely
                    8 -> View your transactions
                    9 -> View account balances\n"""))

                # Check if the user input is correct. If it is not, raise a ValueError
                if response < 1 or response > 9:
                    raise ValueError

                break
            # Dislay to the user the error
            except ValueError:
                print(
                    "Invalid choice! Please type in a valid response: an integer from 1-8 based on your choice of operation. \n")

        # calling a method to record expenses while user inputted 1
        if response == 1:
            recordExpense()
            wb_obj.save("user_db.xlsx")

        # calling a method to record income while user inputted 2
        elif response == 2:
            recordIncome()

        # calling a method to transfer money while user inputted 3
        elif response == 3:
            transferMoney()

        # calling a method to get the total amount spent on a specific category while user inputted 4
        elif response == 4:
            Category_total()

        # calling a method to print the amount spent on a specific day while user inputted 5
        elif response == 5:
            totalexpensesperday()

        # calling a method to sum up all expenses while user inputted 7
        elif response == 6:
            Recorded_expenses()

        # calling a method to print the tip for spending wisely while user inputted 8
        elif response == 7:
            tips()

        # Calling a method to view all transactions at a glance
        elif response == 8:
            viewTransactions()

        elif response == 9:
            viewAccountBalances()

        # While loop to check and ensure that the user input is valid and correct
        while True:
            # Find out if the user would like to perform another operation
            try:
                wb_obj.save("user_db.xlsx")
                tryagain = int(input("""Would you like to perform another operation? 
                                        1 -> Yes please
                                        0 -> No thank you."""))
                break
            # Handle Value Error exception, incase the user inputs a wrong answer
            except ValueError:
                print("Invalid choice. Please insert a number representing your choice")

    print_slow("Here's to making better financial decisions. Good bye :) ")

    # Save the changes made to the excel sheet database
    wb_obj.save("user_db.xlsx")